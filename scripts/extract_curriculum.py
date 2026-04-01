# -*- coding: utf-8 -*-
"""Extract text from curriculum xlsx/hwpx on Desktop for app authoring."""
import re
import zipfile
from pathlib import Path

import openpyxl

DESKTOP = Path(r"C:\Users\SEO_ES\Desktop")
OUT = Path(__file__).resolve().parent.parent / "_extract_curriculum.txt"


def find_materials_dir() -> Path | None:
    for p in DESKTOP.iterdir():
        if not p.is_dir():
            continue
        if any(
            f.suffix.lower() == ".xlsx" and ("5.28" in f.name or "2025" in f.name)
            for f in p.glob("*.xlsx")
        ):
            return p
    for p in DESKTOP.rglob("*.xlsx"):
        if "5.28" in p.name:
            return p.parent
    return None


def dump_xlsx(path: Path, lines: list[str]) -> None:
    lines.append("\n=== XLSX: " + path.name + " ===\n")
    wb = openpyxl.load_workbook(path, data_only=True)
    for sn in wb.sheetnames:
        ws = wb[sn]
        lines.append(f"\n--- Sheet: {sn} ---\n")
        buf: list[str] = []
        for row in ws.iter_rows(values_only=True):
            buf.append("\t".join("" if v is None else str(v).strip() for v in row))
        lines.append("\n".join(buf))


def dump_hwpx(path: Path, lines: list[str]) -> None:
    lines.append("\n=== HWPX: " + path.name + " ===\n")
    with zipfile.ZipFile(path, "r") as z:
        for n in sorted(z.namelist()):
            if not n.lower().endswith(".xml"):
                continue
            low = n.lower()
            if not any(k in low for k in ("section", "content", "body", "header")):
                continue
            raw = z.read(n).decode("utf-8", errors="ignore")
            text = re.sub(r"<[^>]+>", " ", raw)
            text = re.sub(r"\s+", " ", text).strip()
            if len(text) < 30:
                continue
            lines.append(f"\n--- {n} ---\n")
            lines.append(text[:80000])


def main() -> None:
    d = find_materials_dir()
    if d is None:
        OUT.write_text("Could not find materials folder on Desktop", encoding="utf-8")
        print("no dir")
        return
    lines: list[str] = [f"DIR={d}\n"]
    for xf in sorted(d.glob("*.xlsx")):
        try:
            dump_xlsx(xf, lines)
        except Exception as e:
            lines.append(f"XLSX ERROR {xf.name}: {e}\n")
    for hf in sorted(d.glob("*.hwpx")):
        try:
            dump_hwpx(hf, lines)
        except Exception as e:
            lines.append(f"HWPX ERROR {hf.name}: {e}\n")
    OUT.write_text("\n".join(lines), encoding="utf-8")
    print("Wrote", OUT, OUT.stat().st_size)


if __name__ == "__main__":
    main()
